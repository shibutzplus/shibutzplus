import openpyxl
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def convert_docx_tables_to_excel(docx_path, excel_path, sheet_title='מערכות'):
  doc = Document(docx_path)

  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = sheet_title
  ws.sheet_view.rightToLeft = True

  # עיצובים
  title_font = Font(name='Aptos', size=13, bold=True, color='1F497D')
  title_fill = PatternFill(
      start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'
  )
  header_font = Font(name='Aptos', size=10, bold=True, color='000000')
  header_fill = PatternFill(
      start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
  )
  body_font = Font(name='Aptos', size=10)

  thin_border = Border(
      left=Side(style='thin', color='D9D9D9'),
      right=Side(style='thin', color='D9D9D9'),
      top=Side(style='thin', color='D9D9D9'),
      bottom=Side(style='thin', color='D9D9D9'),
  )

  # 1. איסוף כל הפסקאות והטבלאות
  elements = []
  for el in doc.element.body:
    if el.tag.endswith('p'):
      p = Paragraph(el, doc)
      text = p.text.strip()
      if text:
        elements.append(('p', text))
    elif el.tag.endswith('tbl'):
      elements.append(('tbl', Table(el, doc)))

  # 2. התאמת כותרות וטבלאות (כולל מקרים של T P P T)
  matched_pairs = []
  i = 0
  while i < len(elements):
    # אם המסמך מתחיל מטבלה והכותרת אחריה
    if i == 0 and elements[0][0] == 'tbl':
      matched_pairs.append((elements[1][1], elements[0][1]))
      i = 2
    # כותרת ואחריה טבלה
    elif (
        elements[i][0] == 'p'
        and i + 1 < len(elements)
        and elements[i + 1][0] == 'tbl'
    ):
      matched_pairs.append((elements[i][1], elements[i + 1][1]))
      i += 2
    # טבלה ואחריה כותרת
    elif (
        elements[i][0] == 'tbl'
        and i + 1 < len(elements)
        and elements[i + 1][0] == 'p'
    ):
      matched_pairs.append((elements[i + 1][1], elements[i][1]))
      i += 2
    else:
      i += 1

  # 3. כתיבה ל-Excel
  current_row = 1
  for title, tbl in matched_pairs:
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=7,
    )
    current_row += 1

    for r_idx, row in enumerate(tbl.rows):
      reversed_cells = list(reversed(row.cells))
      for c_idx, cell in enumerate(reversed_cells):
        cell_text = cell.text.strip()
        target_cell = ws.cell(
            row=current_row, column=c_idx + 1, value=cell_text
        )

        if r_idx == 0:
          target_cell.font = header_font
          target_cell.fill = header_fill
        else:
          target_cell.font = body_font

        target_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        target_cell.border = thin_border
      current_row += 1

    current_row += 2

  # התאמת רוחב עמודות
  for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
      if cell.value:
        lines = str(cell.value).split('\n')
        longest = max(len(l) for l in lines)
        if longest > max_len:
          max_len = longest
    ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

  wb.save(excel_path)
  print(f'הקובץ נוצר בהצלחה: {excel_path}')


if __name__ == '__main__':
  # הפעלה על קובץ כיתות
  convert_docx_tables_to_excel(
      'כיתות טבלה.docx', 'כיתות.xlsx', sheet_title='מערכות כיתות'
  )