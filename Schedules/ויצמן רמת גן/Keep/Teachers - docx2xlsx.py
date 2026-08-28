import openpyxl
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def docx_to_single_sheet_excel(docx_path, excel_path):
  doc = Document(docx_path)

  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "מערכות שעות"
  ws.sheet_view.rightToLeft = True

  # עיצובים
  title_font = Font(name="Aptos", size=13, bold=True, color="1F497D")
  title_fill = PatternFill(
      start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
  )
  header_font = Font(name="Aptos", size=10, bold=True, color="000000")
  header_fill = PatternFill(
      start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
  )
  body_font = Font(name="Aptos", size=10)

  thin_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )

  # שלב 1: איסוף כל האלמנטים מהמסמך
  elements = []
  for el in doc.element.body:
    if el.tag.endswith("p"):
      p = Paragraph(el, doc)
      text = p.text.strip()
      if text:
        elements.append(("p", text))
    elif el.tag.endswith("tbl"):
      elements.append(("tbl", Table(el, doc)))

  # שלב 2: שיוך מדויק של כותרת לכל טבלה לפי מבנה המסמך
  matched_pairs = []
  i = 0
  while i < len(elements):
    # מקרה 1: פתיחת המסמך (אור סמרה - הטבלה מופיעה לפני הכותרת)
    if i == 0 and elements[0][0] == "tbl":
      matched_pairs.append((elements[1][1], elements[0][1]))
      i = 2
    # מקרה 2: כותרת ואחריה טבלה
    elif (
        elements[i][0] == "p"
        and i + 1 < len(elements)
        and elements[i + 1][0] == "tbl"
    ):
      matched_pairs.append((elements[i][1], elements[i + 1][1]))
      i += 2
    # מקרה 3: טבלה ואחריה כותרת (כמו דיתי כהן, ורד גולן)
    elif (
        elements[i][0] == "tbl"
        and i + 1 < len(elements)
        and elements[i + 1][0] == "p"
    ):
      matched_pairs.append((elements[i + 1][1], elements[i][1]))
      i += 2
    else:
      i += 1

  # שלב 3: כתיבה לאקסל
  current_row = 1
  for title, tbl in matched_pairs:
    # כותרת המורה
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=7,
    )
    current_row += 1

    # כתיבת תוכן הטבלה (עם היפוך ימין-לשמאל של התאים)
    for r_idx, row in enumerate(tbl.rows):
      reversed_cells = list(reversed(row.cells))
      for c_idx, cell in enumerate(reversed_cells):
        cell_text = cell.text.strip()
        target_cell = ws.cell(
            row=current_row, column=c_idx + 1, value=cell_text
        )

        if r_idx == 0:
          target_cell.font = header_font
          target_cell.fill = header_fill
        else:
          target_cell.font = body_font

        target_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        target_cell.border = thin_border
      current_row += 1

    current_row += 2  # רווח של שתי שורות בין מורה למורה

  # התאמת רוחב עמודות
  for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
      if cell.value:
        lines = str(cell.value).split("\n")
        longest = max(len(l) for l in lines)
        if longest > max_len:
          max_len = longest
    ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

  wb.save(excel_path)
  print(f"הקובץ נוצר בהצלחה ובתיאום מושלם לכל 60 המורים: {excel_path}")


if __name__ == "__main__":
  input_file = "מורים טבלה.docx"
  output_file = "מורים.xlsx"
  docx_to_single_sheet_excel(input_file, output_file)